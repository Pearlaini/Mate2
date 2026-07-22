import json
from pathlib import Path
import openpyxl

paths = [p for p in Path(r"C:/Users/Aini/Downloads").glob("QA260529*.xlsx") if "행분리" in p.stem or "_" in p.stem]
# find 행분리 file
split_files = list(Path(r"C:/Users/Aini/Downloads").glob("QA260529_*분리*.xlsx"))
if not split_files:
    split_files = sorted(Path(r"C:/Users/Aini/Downloads").glob("QA260529*.xlsx"), key=lambda p: p.stat().st_mtime)[-1:]
path = split_files[0] if split_files else None
report = {"path": str(path)}
if path and path.exists():
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    report["rows"] = ws.max_row
    report["samples"] = []
    for r in range(2, min(20, ws.max_row + 1)):
        report["samples"].append({
            "r": r,
            "A": ws.cell(r, 1).value,
            "E": ws.cell(r, 5).value,
            "G": str(ws.cell(r, 7).value or "")[:70],
        })
Path(r"d:/py3/_row_split_verify.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
